import os
import json
import torch
from PIL import Image
from torchvision import transforms
from sklearn.metrics import precision_score, recall_score, f1_score, confusion_matrix, roc_auc_score, cohen_kappa_score, \
    classification_report
import matplotlib.pyplot as plt
import numpy as np
import itertools
from openpyxl import load_workbook

from model import GoogLeNet


def plot_confusion_matrix(cm, classes,
                          normalize=False,
                          title='Confusion matrix',
                          cmap=plt.cm.Blues):
    if normalize:
        cm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
        print("Normalized confusion matrix")
    else:
        print('Confusion matrix, without normalization')

    plt.imshow(cm, interpolation='nearest', cmap=cmap)
    plt.title(title)
    plt.colorbar()
    tick_marks = np.arange(len(classes))
    plt.xticks(tick_marks, classes, rotation=45)
    plt.yticks(tick_marks, classes)

    fmt = '.2f' if normalize else 'd'
    thresh = cm.max() / 2.
    for i, j in itertools.product(range(cm.shape[0]), range(cm.shape[1])):
        plt.text(j, i, format(cm[i, j], fmt),
                 horizontalalignment="center",
                 color="white" if cm[i, j] > thresh else "black")

    plt.tight_layout()
    plt.ylabel('True label')
    plt.xlabel('Predicted label')


def append_to_excel(file_path, data):
    if not os.path.exists(file_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Metrics"
        headers = ["Accuracy",
                   "Recall_0", "Recall_1", "Recall_2", "Recall_3", "Recall_4",
                   "Precision_0", "Precision_1", "Precision_2", "Precision_3", "Precision_4",
                   "Macro_Averaged_Precision", "Macro_Averaged_Recall"]
        ws.append(headers)
        wb.save(file_path)

    wb = load_workbook(file_path)
    ws = wb.active

    ws.append(data)

    wb.save(file_path)


def main():
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")

    data_transform = transforms.Compose(
        [transforms.Resize(256),
         transforms.CenterCrop(224),
         transforms.ToTensor(),
         transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])])

    # load image
    imgs_root = r'E:\xwc\pycharm_virtual_cnn\SCUT-FBP5500_1\SCUT-FBP5500\test'
    assert os.path.exists(imgs_root), f"file: '{imgs_root}' dose not exist."

    img_path_list = []
    class_list = {}

    subdirs = [f.name for f in os.scandir(imgs_root) if f.is_dir()]

    for subdir in subdirs:
        subdir_path = os.path.join(imgs_root, subdir)
        images = [os.path.join(subdir_path, i) for i in os.listdir(subdir_path) if i.endswith(".jpg")]
        img_path_list.extend(images)
        class_list[subdir] = images

    json_path = r'E:\xwc\pycharm_virtual_cnn\deep-learning-for-image-processing-master' \
                r'\pytorch_classification\Test5_resnet\class_indices.json'
    assert os.path.exists(json_path), f"file: '{json_path}' dose not exist."

    with open(json_path, "r") as json_file:
        class_indict = json.load(json_file)

    model = GoogLeNet(num_classes=5, aux_logits=False).to(device)

    weights_path = '.\FBPgoogleNet012.pth'
    # weights_path = r'E:\xwc\pycharm_virtual_cnn\SCUT-FBP5500_1\SCUT-FBP5500\FBPgoogleNet.pth'

    assert os.path.exists(weights_path), f"file: '{weights_path}' dose not exist."
    missing_keys, unexpected_keys = model.load_state_dict(torch.load(weights_path, map_location=device), strict=False)

    model.eval()
    batch_size = 8
    correct_count = 0

    all_labels = []
    all_preds = []
    all_probs = []

    with torch.no_grad():
        for ids in range(0, len(img_path_list) // batch_size):
            img_list = []
            for img_path in img_path_list[ids * batch_size: (ids + 1) * batch_size]:
                assert os.path.exists(img_path), f"file: '{img_path}' dose not exist."
                img = Image.open(img_path)
                img = data_transform(img)
                img_list.append(img)

            batch_img = torch.stack(img_list, dim=0)
            output = model(batch_img.to(device)).cpu()
            predict = torch.softmax(output, dim=1)
            probs, classes = torch.max(predict, dim=1)

            for idx, (pro, cla) in enumerate(zip(probs, classes)):
                img_path = img_path_list[ids * batch_size + idx]
                actual_class = None
                for class_name, images in class_list.items():
                    if img_path in images:
                        actual_class = class_name
                        break
                if actual_class and class_indict[str(cla.numpy())] == actual_class:
                    correct_count += 1
                    correct_str = "正确"
                else:
                    correct_str = "错误"

                all_labels.append(actual_class)
                all_preds.append(class_indict[str(cla.numpy())])
                all_probs.append(pro.numpy())

                print("image: {}  预测类别: {}  实际类别: {}  最大概率: {:.3}  预测是否正确: {}".format(
                    img_path, class_indict[str(cla.numpy())], actual_class, pro.numpy(), correct_str
                ))

        macro_precision = precision_score(all_labels, all_preds, average='macro', zero_division=0)
        macro_recall = recall_score(all_labels, all_preds, average='macro', zero_division=0)
        macro_f1 = f1_score(all_labels, all_preds, average='macro', zero_division=0)
        micro_precision = precision_score(all_labels, all_preds, average='micro', zero_division=0)
        micro_recall = recall_score(all_labels, all_preds, average='micro', zero_division=0)
        micro_f1 = f1_score(all_labels, all_preds, average='micro', zero_division=0)
        kappa = cohen_kappa_score(all_labels, all_preds)

        print("\n宏观平均精确率（Macro-Averaged Precision）: {:.4f}".format(macro_precision))
        print("宏观平均召回率（Macro-Averaged Recall）: {:.4f}".format(macro_recall))
        print("宏观平均 F1 分数（Macro-Averaged F1-score）: {:.4f}".format(macro_f1))
        print("微观平均精确率（Micro-Averaged Precision）: {:.4f}".format(micro_precision))
        print("微观平均召回率（Micro-Averaged Recall）: {:.4f}".format(micro_recall))
        print("微观平均 F1 分数（Micro-Averaged F1-score）: {:.4f}".format(micro_f1))
        print("Kappa 系数: {:.4f}".format(kappa))

        cm = confusion_matrix(all_labels, all_preds, labels=list(class_indict.values()))
        print("混淆矩阵:\n", cm)

        plt.figure()
        plot_confusion_matrix(cm, classes=list(class_indict.values()), title='Confusion Matrix')
        # plt.show()

        if len(np.unique(all_labels)) == 2:  # 只有在二分类时才计算ROC AUC
            roc_auc = roc_auc_score(all_labels, all_probs)
            print("ROC AUC: {:.4f}".format(roc_auc))

        print("\n每类的精确率、召回率和 F1 分数:")
        report = classification_report(all_labels, all_preds, target_names=list(class_indict.values()), zero_division=0)
        print(report)

        report_dict = classification_report(all_labels, all_preds, target_names=list(class_indict.values()),
                                            zero_division=0, output_dict=True)

        recalls = []
        precisions = []
        for class_name, metrics in report_dict.items():
            if isinstance(metrics, dict):
                recalls.append(metrics['recall'])
                precisions.append(metrics['precision'])
                print(f"类 '{class_name}' 的精确率: {metrics['precision']:.4f}")
                print(f"类 '{class_name}' 的召回率: {metrics['recall']:.4f}")

        accuracy = correct_count / len(img_path_list)
        data_to_append = [accuracy] + recalls + precisions + [macro_precision, macro_recall]

        excel_file_path = 'TestResult.xlsx'
        append_to_excel(excel_file_path, data_to_append)

        print("\n总测试数量有：{}，总的预测正确数量: {},准确率为：{}".format(len(img_path_list), correct_count, accuracy))


if __name__ == '__main__':
    main()
